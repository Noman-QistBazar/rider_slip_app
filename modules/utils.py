import os
import hashlib
import pandas as pd
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime, timedelta
from gspread_dataframe import set_with_dataframe
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json

BRANCH_DATA_FILE = "branch_data.json"


# Load branch data from file
def load_branch_data():
    if os.path.exists(BRANCH_DATA_FILE):
        with open(BRANCH_DATA_FILE, "r") as f:
            return json.load(f)
    return {}

# Save branch data to file
def save_branch_data(data):
    with open(BRANCH_DATA_FILE, "w") as f:
        json.dump(data, f)
        
# Constants
DATA_FILE = "data/all_branch_data.xlsx"
REQUESTS_FILE = "data/requests.csv"
BACKUP_FOLDER = "data/backups"
GOOGLE_SHEET_NAME = "Rider Slip Data"  # ✅ Update to match your actual Google Sheet name

# 🎨 Optional animated UI component
def show_animated_submit_button():
    import streamlit as st
    st.markdown("""
        <style>
        .fade-in {
            animation: fadeIn 1s ease-in-out;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        </style>
    """, unsafe_allow_html=True)

    submitter_name = st.text_input("Enter Your Name")

    if submitter_name.strip():
        st.markdown('<div class="fade-in">', unsafe_allow_html=True)
        if st.button("🚀 Submit Slip"):
            st.success(f"Submitted by {submitter_name}")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.warning("Please enter your name to proceed.")

# ✅ Ensure valid Excel structure
def ensure_valid_excel(file_path=DATA_FILE):
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        wb = Workbook()
        wb.save(file_path)
    else:
        try:
            with open(file_path, 'rb') as f:
                load_workbook(filename=BytesIO(f.read()))
        except:
            wb = Workbook()
            wb.save(file_path)

# 🛡️ Create backup of the Excel file
def backup_data():
    today = datetime.today().strftime('%Y-%m-%d')
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    backup_file = f"{BACKUP_FOLDER}/backup_{today}.xlsx"
    if os.path.exists(DATA_FILE):
        from shutil import copyfile
        copyfile(DATA_FILE, backup_file)

# 🔐 File integrity check
def file_hash(file):
    return hashlib.sha256(file.getbuffer()).hexdigest()

# 🚫 Detect duplicate slip images
def is_duplicate_image(hash_val, folder_path):
    hash_file = os.path.join(folder_path, "hashes.txt")
    if not os.path.exists(hash_file):
        return False
    with open(hash_file, "r") as f:
        return hash_val in set(f.read().splitlines())

# 💾 Save image and record its hash
def save_image_and_hash(file, folder_path, filename):
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    with open(file_path, "wb") as f:
        f.write(file.getbuffer())
    hash_val = file_hash(file)
    hash_file = os.path.join(folder_path, "hashes.txt")
    with open(hash_file, "a") as f:
        f.write(hash_val + "\n")

# 💰 Commission calculation
def calculate_commission(slip_qty, slip_type):
    return slip_qty * 50 if slip_type == "Online Slip" else 25

# 📅 Generate recent weekly ranges
def generate_weeks(n=4):
    today = datetime.today()
    return [(today - timedelta(days=today.weekday() + i * 7),
             today - timedelta(days=today.weekday() + i * 7) + timedelta(days=6))
            for i in range(n)]

# 🔄 Sync to Google Sheets
def save_to_google_sheets():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("credentials/google-credentials.json", scope)
    client = gspread.authorize(creds)

    if not os.path.exists(DATA_FILE):
        return

    spreadsheet = client.open(GOOGLE_SHEET_NAME)

    with pd.ExcelFile(DATA_FILE) as xls:
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)

            try:
                worksheet = spreadsheet.worksheet(sheet_name)
                worksheet.clear()
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols="20")

            set_with_dataframe(worksheet, df)
